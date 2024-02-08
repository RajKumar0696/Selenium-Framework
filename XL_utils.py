import openpyxl
from openpyxl.styles import PatternFill
path = "Test_data/login_test_data.xlsx"


def get_row_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def get_column_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def read_data(file, sheetName, rowNum,columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    value = sheet.cell(rowNum, columnNum).value
    return value


def write_data(file, sheetName, rowNum, columnNum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNum, columnNum).value = data
    workbook.save(file)


def fill_green_color(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    green = PatternFill(start_color="0DF955", end_color="0DF955", fill_type="solid")
    sheet.cell(rowNum, columnNum).fill = green
    workbook.save(file)


def fill_red_color(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    red = PatternFill(start_color="F9180D", end_color="F9180D", fill_type="solid")
    sheet.cell(rowNum, columnNum).fill = red
    workbook.save(file)
