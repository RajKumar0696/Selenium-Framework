import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowNum, columnNum).value


def writeData(file, sheetName, rowNum, columnNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNum, columnNum).value = data
    workbook.save(file)


def fill_green_color(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    green = PatternFill(start_color="0DF955", end_color="0DF955", fill_type="solid")
    sheet.cell(rowNum, columnNum).fill = green
    workbook.save(file)


def fill_red_color(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    red = PatternFill(start_color="F9180D", end_color="F9180D", fill_type="solid")
    sheet.cell(rowNum, columnNum).fill = red
    workbook.save(file)
